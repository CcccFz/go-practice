# -*- coding: utf-8 -*-
__author__ = 'Duome'


from openpyxl import load_workbook, Workbook, compat


class Xlsx(object):

    def __init__(self, filename):
        """ book_data = {
                'sheetname': [[rowdata], [rowdata], ...],
                'sheetname': [[rowdata], [rowdata], ...],
                '': [[], [], ...],
                ...
                }
        """
        self.book_data = {}
        self.book = load_workbook(filename=filename)
        self.sheetnames = self.book.get_sheet_names()
        self.head = []


    def get_data(self):
        """ 获取excel数据 """
        for sheetname in self.sheetnames:
            self.book_data[sheetname] = []
            sheet = self.book.get_sheet_by_name(sheetname)

            row_nums = len(sheet['A'])
            col_nums = len(sheet['1'])

            for row in compat.range(row_nums):
                self.book_data[sheetname].append([])
                for col in compat.range(col_nums):
                    cell_value = sheet.cell(row=row+1, column=col+1).value

                    if row == 0:
                        self.head.append(cell_value)
                    self.book_data[sheetname][row].append(cell_value)

        return self.book_data


    def save_data(self, book_data, filename):
        """ 保存excel数据 """
        new_book = Workbook()
        new_book.remove_sheet(new_book.active)

        for sheetname in self.sheetnames:
            sheet = new_book.create_sheet(title='%s' % sheetname)
            sheet_data = book_data[sheetname]
            for row in sheet_data:
                sheet.append(row)

        new_book.save(filename=filename)
